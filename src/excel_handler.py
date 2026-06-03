import enum
from os import path
import os
import openpyxl
import pandas as pd

from src import config

class DataSetType(enum.Enum):
    RAW = "raw"
    PROCESSED = "processed"

class ExcelHandler:    
    cfg: config.Excel
    raw_workbook: openpyxl.workbook.workbook.Workbook
    processed_workbook: openpyxl.workbook.workbook.Workbook
    df: pd.DataFrame
    
    def __init__(self, cfg: config.Excel):
        self.cfg = cfg
        self.processed_workbook = openpyxl.load_workbook(self.cfg.processed_path)
        
    def save(self, type: DataSetType, df: pd.DataFrame) -> None:
        match type:
            case DataSetType.RAW:
                df.to_excel(
                    self.cfg.raw_path,
                    sheet_name=self.cfg.raw_sheet_name,
                    index=False,
                    engine='openpyxl',
                )
            case DataSetType.PROCESSED:
                df.to_excel(
                    self.cfg.processed_path,
                    sheet_name=self.cfg.processed_sheet_name,
                    index=False,
                    engine='openpyxl',
                )
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")
                
    def get_df(self, type: DataSetType) -> pd.DataFrame:
        match type:
            case DataSetType.RAW:
                data = []
                raw_dir = path.dirname(self.cfg.raw_path)  # "data/raw"

                for i in os.listdir(path.dirname(self.cfg.raw_path)):
                    if i.endswith(".csv"):
                        df = pd.read_csv(path.join(raw_dir, i))
                        data.append(df)
                        
                return data
            
            case DataSetType.PROCESSED:
                return pd.read_excel(self.cfg.processed_path, sheet_name=self.cfg.processed_sheet_name)
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")

    def clear(self, type: DataSetType) -> None:
        match type:
            case DataSetType.RAW:
                self.raw_workbook[self.cfg.raw_sheet_name].delete_rows(1, self.raw_workbook[self.cfg.raw_sheet_name].max_row)
                self.raw_workbook.save(self.cfg.raw_path)
            case DataSetType.PROCESSED:
                self.processed_workbook[self.cfg.processed_sheet_name].delete_rows(1, self.processed_workbook[self.cfg.processed_sheet_name].max_row)
                self.processed_workbook.save(self.cfg.processed_path)
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")
                  
    def set_raw_path(self, raw_path: str) -> None:
        self.raw_workbook = openpyxl.load_workbook(raw_path)
        self.cfg.raw_path = raw_path